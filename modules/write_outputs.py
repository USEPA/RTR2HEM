import os, shutil
import msaccessdb, pypyodbc
import pandas as pd
from openpyxl import load_workbook
from modules.outputs_writer.emissions_loc import EmissionLoc
from modules.outputs_writer.fac_address import FacilityAddress
from modules.outputs_writer.fac_list import FacilityList
from modules.outputs_writer.hap_emissions import HapEmissions
from modules.utils import src_cat_name, timestamp, emission_type, only_category

"""
If both category and wholesale selected then need to produce two sets of files for each template
"""


class WriteOuputs:
    templates_fp = "templates"
    filename_base = f"{src_cat_name}{timestamp}_{emission_type.split(' ')[0]}"

    def __init__(self, df):
        self.df = df

        self.out_fp = os.path.join(
            "outputs", f"{self.filename_base}_HEMInputsAndXWalks"
        )
        self.accdb_fp = os.path.join(self.out_fp, f"{self.filename_base}_XWalks.accdb")
        self.odbc_conn_str = ""

        self.create_folder()

    def create_folder(self):
        if not os.path.exists("outputs"):
            os.mkdir("outputs")
        if os.path.exists(self.out_fp):
            shutil.rmtree(self.out_fp)
        os.mkdir(self.out_fp)

    def create_accdb(self):
        msaccessdb.create(self.accdb_fp)
        self.odbc_conn_str = (
            r"Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ="
            + self.accdb_fp
            + ";"
        )

    def run(self):
        emis_loc = EmissionLoc(self.df).create()
        self.write_to_template(emis_loc)

        fac_address = FacilityAddress(self.df).create()
        self.write_to_template(fac_address)

        fac_list = FacilityList(self.df).create()
        self.write_to_template(fac_list)

        hap_emissions = HapEmissions(self.df).create()
        self.write_to_template(hap_emissions)

        self.write_access_history()

    def write_access_history(self):
        self.create_accdb()
        conn = pypyodbc.connect(self.odbc_conn_str)

        accdb = conn.cursor()
        accdb.execute(
            """CREATE TABLE Table1 (
                        ID autoincrement,
                        Col1 varchar(50),
                        Col2 double,
                        Col3 datetime);"""
        )
        accdb.commit()

    def write_to_template(self, result):
        template_name = result.template_name
        sheetname = result.sheet_name
        row = result.rowstart

        template_src = os.path.join(self.templates_fp, f"{template_name}.xlsx")
        filename = f"{self.filename_base}_{result.filename}_Cat"
        output_dst = os.path.join(self.out_fp, f"{filename}.xlsx")

        # Write category records
        shutil.copyfile(template_src, output_dst)

        writer = pd.ExcelWriter(
            output_dst, engine="openpyxl", mode="a", if_sheet_exists="overlay"
        )
        result.cat_df.to_excel(
            writer,
            sheet_name=sheetname,
            header=False,
            index=False,
            startrow=row,
        )
        writer.close()
        # self.insert_notes(template_src, output_dst)

        # Write whole records
        if not only_category:
            filename = f"{self.filename_base}_{result.filename}_Whole"
            output_dst = os.path.join(self.out_fp, f"{filename}.xlsx")
            shutil.copyfile(template_src, output_dst)

            writer = pd.ExcelWriter(
                output_dst, engine="openpyxl", mode="a", if_sheet_exists="overlay"
            )
            result.whole_df.to_excel(
                writer,
                sheet_name=sheetname,
                header=False,
                index=False,
                startrow=row,
            )
            writer.close()

    def insert_notes(self, src_fp, dst_fp):
        template_wb = load_workbook(src_fp)
        template_ws = template_wb.active

        out_wb = load_workbook(dst_fp)
        out_ws = out_wb.active

        for i, row in enumerate(template_ws):
            if i > 2:
                break
            for cell in row:
                if cell.comment:
                    out_ws[cell.coordinate].comment = cell.comment

        out_wb.save(dst_fp)
        out_wb.close()
        template_wb.close()
